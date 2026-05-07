#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
LINE日報自動登録システム
LINEグループからの日報メッセージを自動的にExcelファイルに反映
"""

from flask import Flask, request, abort
from linebot import LineBotApi, WebhookHandler
from linebot.exceptions import InvalidSignatureError
from linebot.models import MessageEvent, TextMessage, TextSendMessage
import openpyxl
from datetime import datetime
import os
import re
from dotenv import load_dotenv
import logging

# 環境変数の読み込み
load_dotenv()

# ロギング設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# LINE API設定
CHANNEL_ACCESS_TOKEN = os.getenv('LINE_CHANNEL_ACCESS_TOKEN', 'YOUR_CHANNEL_ACCESS_TOKEN')
CHANNEL_SECRET = os.getenv('LINE_CHANNEL_SECRET', 'YOUR_CHANNEL_SECRET')

line_bot_api = LineBotApi(CHANNEL_ACCESS_TOKEN)
handler = WebhookHandler(CHANNEL_SECRET)

# Excelファイルのパス
EXCEL_FILE = os.getenv('EXCEL_FILE_PATH', 'LINK_工数表_5月分.xlsx')


class DailyReportParser:
    """日報メッセージのパーサー"""
    
    @staticmethod
    def parse_message(message_text):
        """
        LINEメッセージから日報情報を抽出
        
        フォーマット例:
        5/4
        鹿島石油
        加藤、金澤、ニカジ
        佐々木、アリ、ダニ
        6名
        """
        lines = message_text.strip().split('\n')
        
        if len(lines) < 3:
            return None
        
        data = {
            'date': None,
            'site_name': None,
            'workers': [],
            'total_count': 0
        }
        
        # 日付解析 (例: "5/4", "5/5")
        date_match = re.match(r'(\d{1,2})/(\d{1,2})', lines[0].strip())
        if date_match:
            month = int(date_match.group(1))
            day = int(date_match.group(2))
            current_year = datetime.now().year
            data['date'] = datetime(current_year, month, day)
        else:
            return None
        
        # 現場名 (2行目)
        data['site_name'] = lines[1].strip()
        
        # 作業員名の抽出 (3行目以降、"X名"の前まで)
        workers = []
        for line in lines[2:]:
            line = line.strip()
            
            # "X名"のパターンをチェック
            count_match = re.search(r'(\d+)\s*名', line)
            if count_match:
                data['total_count'] = int(count_match.group(1))
                break
            
            # 作業員名を抽出（カンマまたは、で区切られている）
            # 全角・半角スペースも考慮
            if line:
                # カンマ区切りまたは、区切りで分割
                names = re.split(r'[,、]', line)
                workers.extend([name.strip() for name in names if name.strip()])
        
        data['workers'] = workers
        
        # 作業員数が抽出できていない場合、作業員名の数をカウント
        if data['total_count'] == 0:
            data['total_count'] = len(workers)
        
        logger.info(f"パース結果: {data}")
        return data


class ExcelUpdater:
    """Excelファイル更新クラス"""
    
    def __init__(self, excel_path):
        self.excel_path = excel_path
        
    def update_daily_report(self, report_data, company_name='雅工業'):
        """
        日報データをExcelに反映
        
        Args:
            report_data: パースされた日報データ
            company_name: 会社名（シート名）
        """
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            
            # 会社名のシートを取得
            if company_name not in wb.sheetnames:
                logger.error(f"シート '{company_name}' が見つかりません")
                return False
            
            ws = wb[company_name]
            
            # 日付に対応する行を見つける
            target_date = report_data['date']
            target_row = self._find_date_row(ws, target_date)
            
            if target_row is None:
                logger.error(f"日付 {target_date.strftime('%Y/%m/%d')} に対応する行が見つかりません")
                return False
            
            # データを書き込み
            # C列: 現場名
            ws.cell(row=target_row, column=3).value = report_data['site_name']
            
            # D列: 工数（作業員数）
            ws.cell(row=target_row, column=4).value = report_data['total_count']
            
            # 保存
            wb.save(self.excel_path)
            logger.info(f"Excel更新成功: {target_date.strftime('%m/%d')} - {report_data['site_name']} - {report_data['total_count']}名")
            
            return True
            
        except Exception as e:
            logger.error(f"Excel更新エラー: {str(e)}", exc_info=True)
            return False
    
    def _find_date_row(self, worksheet, target_date):
        """
        指定された日付に対応する行番号を見つける
        
        Args:
            worksheet: openpyxlのワークシート
            target_date: datetime オブジェクト
        
        Returns:
            行番号 (1-indexed) または None
        """
        # A列（日付列）を走査
        # 注: Excelの数式セルは直接評価できないため、行番号から推測
        # 4行目 = 1日、5行目 = 2日、...の規則を利用
        
        target_day = target_date.day
        target_month = target_date.month
        
        # 行番号 = 3 + target_day（4行目が1日なので）
        estimated_row = 3 + target_day
        
        if estimated_row < 4 or estimated_row > 34:
            logger.error(f"日付が範囲外です: {target_date.strftime('%Y/%m/%d')}")
            return None
        
        logger.info(f"日付 {target_date.strftime('%Y/%m/%d')} → 行 {estimated_row}")
        return estimated_row


# グローバルインスタンス
parser = DailyReportParser()
excel_updater = ExcelUpdater(EXCEL_FILE)


@app.route("https://line-excel-bot.onrender.com/callback", methods=['POST'])
    """LINEからのWebhookを受信"""
    signature = request.headers.get('X-Line-Signature', '')
    body = request.get_data(as_text=True)
    
    logger.info(f"Webhook受信: signature={signature[:10]}...")
    
    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        logger.error("Invalid signature")
        abort(400)
    
    return 'OK'


@handler.add(MessageEvent, message=TextMessage)
def handle_message(event):
    """メッセージイベントの処理"""
    
    message_text = event.message.text
    logger.info(f"メッセージ受信: {message_text[:50]}...")
    
    # 日報フォーマットかどうかをチェック
    # "X/X" で始まるメッセージを日報と判定
    if re.match(r'^\d{1,2}/\d{1,2}', message_text.strip()):
        
        # メッセージをパース
        report_data = parser.parse_message(message_text)
        
        if report_data is None:
            logger.warning("日報のパースに失敗しました")
            # パース失敗の通知は送らない（静かに失敗）
            return
        
        # Excelに反映
        success = excel_updater.update_daily_report(report_data)
        
        if success:
            # 成功通知
            reply_message = (
                f"✅ 日報を登録しました\n"
                f"日付: {report_data['date'].strftime('%m/%d')}\n"
                f"現場: {report_data['site_name']}\n"
                f"工数: {report_data['total_count']}名"
            )
        else:
            # 失敗通知
            reply_message = "❌ 日報の登録に失敗しました"
        
        # 返信（オプション - コメントアウトで無効化可能）
        try:
            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(text=reply_message)
            )
        except Exception as e:
            logger.error(f"返信エラー: {str(e)}")


@app.route("/health", methods=['GET'])
def health_check():
    """ヘルスチェック用エンドポイント"""
    return {
        'status': 'ok',
        'timestamp': datetime.now().isoformat()
    }


if __name__ == "__main__":
    # Excelファイルの存在確認
    if not os.path.exists(EXCEL_FILE):
        logger.warning(f"Excelファイルが見つかりません: {EXCEL_FILE}")
    
    # サーバー起動
    port = int(os.getenv('PORT', 5000))
    logger.info(f"サーバー起動: ポート {port}")
    app.run(host='0.0.0.0', port=port, debug=True)
